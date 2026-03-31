import io
from datetime import datetime

from fastapi import APIRouter, Request, Depends
from fastapi.responses import StreamingResponse, RedirectResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from app.deps import get_current_user
from app.services.supabase_client import get_user_client

router = APIRouter()

HEADERS = ["Nome", "Projeto", "Prioridade", "Status", "Deadline", "Criada em", "Concluída em", "Atualizações"]
HEADER_FILL = PatternFill("solid", fgColor="1a1a2e")
HEADER_FONT = Font(color="FFFFFF", bold=True)
COL_WIDTHS  = [40, 20, 12, 14, 14, 18, 18, 60]


def _fmt_dt(iso_str):
    if not iso_str:
        return ""
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y %H:%M")
    except Exception:
        return iso_str[:16] if len(iso_str) >= 16 else iso_str


def _task_to_row(task):
    updates = sorted(task.get("task_updates") or [], key=lambda u: u.get("created_at", ""))
    updates_text = " | ".join(
        f'[{i+1}] {_fmt_dt(u.get("created_at"))}: {u.get("text","")}'
        for i, u in enumerate(updates)
    )
    priority_map = {"critica": "Crítica", "urgente": "Urgente", "normal": "Normal"}
    return [
        task.get("name", ""),
        task.get("project", "") or "",
        priority_map.get(task.get("priority", ""), task.get("priority", "")),
        task.get("status", ""),
        task.get("deadline", "") or "",
        _fmt_dt(task.get("created_at")),
        _fmt_dt(task.get("completed_at")),
        updates_text,
    ]


def _build_xlsx(tasks: list, sheet_name: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    for col_idx, (header, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = width

    for task in tasks:
        ws.append(_task_to_row(task))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _xlsx_response(data: bytes, filename: str) -> StreamingResponse:
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/export/active")
async def export_active(request: Request, user=Depends(get_current_user)):
    if isinstance(user, RedirectResponse):
        return user
    client = get_user_client(user["access_token"], user["refresh_token"])
    tasks = (
        client.table("tasks")
        .select("*, task_updates(*)")
        .eq("user_id", user["user_id"])
        .eq("status", "active")
        .order("created_at")
        .execute()
        .data or []
    )
    priority_order = {"critica": 0, "urgente": 1, "normal": 2}
    tasks.sort(key=lambda t: priority_order.get(t.get("priority", "normal"), 2))
    from datetime import date
    filename = f"jtasks-ativas-{date.today().isoformat()}.xlsx"
    return _xlsx_response(_build_xlsx(tasks, "Ativas"), filename)


@router.get("/export/completed")
async def export_completed(request: Request, user=Depends(get_current_user)):
    if isinstance(user, RedirectResponse):
        return user
    client = get_user_client(user["access_token"], user["refresh_token"])
    tasks = (
        client.table("tasks")
        .select("*, task_updates(*)")
        .eq("user_id", user["user_id"])
        .eq("status", "completed")
        .order("completed_at", desc=True)
        .execute()
        .data or []
    )
    from datetime import date
    filename = f"jtasks-concluidas-{date.today().isoformat()}.xlsx"
    return _xlsx_response(_build_xlsx(tasks, "Concluídas"), filename)
